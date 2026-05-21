"""
========================================================
  EXPORT places.db → places.xlsx
  ─────────────────────────────────────────────────────
  Usage:
    python to_excel.py                  # default: places.db → places.xlsx
    python to_excel.py -d my.db -o out.xlsx

  Completed rows only.  Each place becomes one row:
    • Basic fields as flat columns
    • Hours:  hours_Mon … hours_Sun
    • Busy hours: peak_Mon … peak_Sun  (e.g. "9 AM (84%)")
    • hours_might_differ as pipe-separated string
========================================================
"""

import argparse
import json
import sqlite3
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependencies — run: pip install pandas openpyxl")


DAYS = ["Sunday", "Monday", "Tuesday", "Wednesday",
        "Thursday", "Friday", "Saturday"]

HEADER_FILL  = "1F4E79"   # dark blue
HEADER_FONT  = "FFFFFF"   # white text
ALT_ROW_FILL = "D6E4F0"   # light blue for alternate rows


# ─── FLATTEN ─────────────────────────────────────────────────────────────────

def _flatten(url: str, db_row: sqlite3.Row, details: dict) -> dict:
    row: dict = {
        "url":                  url,
        "name":                 details.get("name")         or db_row["name"] or "",
        "category":             details.get("category",     ""),
        "rating":               details.get("rating",       ""),
        "review_count":         details.get("review_count", ""),
        "phone":                details.get("phone",        ""),
        "address":              details.get("address",      ""),
        "website":              details.get("website",      ""),
        "booking_url":          details.get("booking_url",  ""),
        "wheelchair_accessible":details.get("wheelchair_accessible", ""),
        "keyword":              db_row["keyword"]      or "",
        "source_city":          db_row["source_url"]   or "",
        "discovered_at":        db_row["discovered_at"] or "",
        "completed_at":         db_row["completed_at"]  or "",
    }

    # Hours: one column per day (Mon–Sun order for readability)
    hours = details.get("hours", {})
    for day in DAYS:
        row[f"hours_{day[:3]}"] = hours.get(day, "")

    row["hours_might_differ"] = " | ".join(details.get("hours_might_differ", []))

    # Busy hours: peak time per day
    busy = details.get("busy_hours", {})
    for day in DAYS:
        day_data = busy.get(day)
        if day_data is None:
            row[f"peak_{day[:3]}"] = "Closed"
        elif isinstance(day_data, dict) and day_data:
            peak_time = max(day_data, key=lambda t: day_data[t])
            row[f"peak_{day[:3]}"] = f"{peak_time} ({day_data[peak_time]}%)"
        else:
            row[f"peak_{day[:3]}"] = ""

    return row


# ─── STYLE ───────────────────────────────────────────────────────────────────

def _style(ws, n_rows: int) -> None:
    """Freeze header, bold + colour header row, auto-width columns, alt rows."""
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color=HEADER_FONT)
    alt_fill    = PatternFill("solid", fgColor=ALT_ROW_FILL)

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=False)

    for row_idx in range(2, n_rows + 2, 2):   # every other data row
        for cell in ws[row_idx]:
            cell.fill = alt_fill

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Export places.db to Excel")
    parser.add_argument("-d", "--db",  default="places.db",    help="SQLite DB path")
    parser.add_argument("-o", "--out", default="places.xlsx",  help="Output .xlsx path")
    args = parser.parse_args()

    db_path = Path(args.db)
    if not db_path.exists():
        sys.exit(f"Database not found: {db_path}")

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT url, name, keyword, source_url, discovered_at,
               completed_at, details_json
          FROM places
         WHERE status = 'completed'
           AND details_json IS NOT NULL
         ORDER BY completed_at DESC
        """
    ).fetchall()
    conn.close()

    if not rows:
        sys.exit("No completed rows in database yet — run pass2.py first.")

    records = []
    for r in rows:
        try:
            details = json.loads(r["details_json"])
        except (json.JSONDecodeError, TypeError):
            details = {}
        records.append(_flatten(r["url"], r, details))

    df = pd.DataFrame(records)

    out_path = Path(args.out)
    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Places")
        _style(writer.sheets["Places"], len(df))

    print(f"Exported {len(df):,} rows → {out_path}")
    print(f"Columns: {len(df.columns)}  "
          f"(basic: 14, hours: {len(DAYS)}, peak: {len(DAYS)}, misc: 1)")


if __name__ == "__main__":
    main()
